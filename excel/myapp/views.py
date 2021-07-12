from time import sleep
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
import socket

from django.shortcuts import render
import openpyxl

def index(request):
    if "GET" == request.method:
        return render(request, 'index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        ws = wb.active
        print(ws)

        # reading a cell
        #print(worksheet["A2"].value)
        driver = webdriver.Chrome('E:\Test\excel\chromedriver.exe')
        driver.get("http://web.whatsapp.com")

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for column in worksheet.iter_rows():
            column_data = list()
            for cell in column:
                column_data.append(str(cell.value))
            #print(column_data[0])
            text = column_data[1]
            Contact = column_data[0]
            
            driver.get("https://web.whatsapp.com/send?phone={}&source=&data=#".format(Contact))
            
            try:
                driver.switch_to_alert().accept()
            except Exception as e:
                pass

            element_present = EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/footer/div[1]/div[2]/div/div[2]'))
            WebDriverWait(driver, 150).until(element_present)
            
            txt_box = driver.find_element(By.XPATH, '//*[@id="main"]/footer/div[1]/div[2]/div/div[2]')
            txt_box.send_keys(text)
            txt_box.send_keys("\n")
    
    return render(request,'index.html')
    #excel_data.append(column_data)